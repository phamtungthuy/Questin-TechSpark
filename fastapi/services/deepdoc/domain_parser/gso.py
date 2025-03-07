import copy
from io import BytesIO
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from services.deepdoc.parser import ExcelParser
from dateutil.parser import parse as datetime_parse
from services.nlp import rag_tokenizer, tokenize
import json
import unicodedata
from utils import get_uuid

def remove_vietnamese_diacritics(text):
    vietnamese_map = {
        'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        'đ': 'd'
    }

    # Normalize text to decomposed form
    text = unicodedata.normalize('NFD', text)
    
    # Remove diacritics
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    
    # Replace Vietnamese characters using mapping
    text = text.lower()
    for vietnamese_char, ascii_char in vietnamese_map.items():
        text = text.replace(vietnamese_char, ascii_char)

     # Remove any remaining non-ASCII characters
    text = ''.join(c for c in text if ord(c) < 128)
    
    return text

def adjust_header(header):
    # Loại bỏ dấu
    if header is None: return None
    normalized_header = remove_vietnamese_diacritics(header)
    # Thêm "_" nếu bắt đầu bằng số
    if normalized_header and normalized_header[0].isdigit():
        normalized_header = "_" + normalized_header
    return normalized_header

class Excel(ExcelParser):
    def __call__(self, fnm, binary=None, from_page=0,
                 to_page=1000000000, callback=None):
        if not binary:
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))
            
        res, fails, done = [], [], 0
        rn = 0
        count = 0
        titles = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows: continue
            missed = []
            multi_header = True if rows[3][0].value is None else False
            
            titles.append(rows[0][0].value)
            headers = []
            if multi_header:
                current_cell_value = ""
                for i, cell in enumerate(rows[2]):
                    if cell.value is not None or rows[3][i].value is not None:
                        if cell.value is not None:
                            current_cell_value = cell.value
                        value = rows[3][i].value
                        headers.append(f"{current_cell_value} {value}")
                    else:
                        cell_values = [row[i].value for row in rows[2:]]
                        column_name = column_name_type(cell_values)
                        stt = sum([1 for header in headers if header.startswith(column_name)])
                        if stt > 0:
                            headers.append(f"{column_name}_{stt + 1}")
                        else:
                            headers.append(f"{column_name}")
            else:
                for i, cell in enumerate(rows[2]):
                    if cell.value is not None:
                        headers.append(cell.value)
                    else:
                        cell_values = [row[i].value for row in rows[2:]]
                        column_name = column_name_type(cell_values)
                        stt = sum([1 for header in headers if header.startswith(column_name)])
                        if stt > 0:
                            headers.append(f"{column_name}_{stt + 1}")
                        else:
                            headers.append(f"{column_name}")
            
            missed = set([i for i, h in enumerate(headers) if h is None])
            headers = [header for header in headers if header is not None]
            if not headers:continue
            data = []
            ii_rows = rows[4:] if multi_header else rows[3:]
            for i, r in enumerate(ii_rows):
                rn += 1
                if rn - 1 < from_page:
                    continue
                if rn - 1 >= to_page:
                    break
                row = [
                    cell.value for ii,
                    cell in enumerate(r) if ii not in missed]

                if len(row) != len(headers):
                    fails.append(str(i))
                    continue
                data.append(row)
                done += 1
            res.append(pd.DataFrame(np.array(data), columns=headers))

        callback(0.3, ("Extract records: {}~{}".format(from_page + 1, min(to_page, from_page + rn)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        return res, titles
    
def trans_datatime(s):
    try:
        return datetime_parse(s.strip()).strftime("%Y-%m-%d %H:%M:%S")
    except Exception as e:
        pass

def trans_bool(s):
    if re.match(r"(đúng|true|yes|是|\*|✓|✔|☑|✅|√)$",
                str(s).strip(), flags=re.IGNORECASE):
        return "yes"
    if re.match(r"(không|sai|false|no|否|⍻|×)$", str(s).strip(), flags=re.IGNORECASE):
        return "no"

def column_name_type(arr):
    arr = list(arr)
    counts = { "category": 0, "year": 0, "location": 0, "indicator": 0, "scope": 0}
    for a in arr:
        if a is None:
            continue
        if re.match(r"[+-]?[0-9.]+$", str(a).replace("%%", "")):
            counts["year"] += 1
        if re.search(r"(thành thị|hà nội|hồ chí minh|an giang|nông thôn|đông nam bộ|tây nguyên|đồng bằng sông hồng|đồng bằng sông cửu long|duyên hải miền trung|trung du và miền núi phía bắc)", str(a), flags=re.IGNORECASE):
            counts["location"] += 1
        if re.search(r"(khai thác|sản xuất|hàng hải quốc tế|than|dầu thô|khai khoáng|sắt, thép|muối biển| nhà máy điện|nhà máy lọc dầu|hàng không quốc tế)", str(a), flags=re.IGNORECASE):
            counts["category"] += 1
        if re.search(r"(tỷ lệ|tiêu thụ|cơ cấu|tỷ suất|tỷ số)", str(a), flags=re.IGNORECASE):
            counts["indicator"] += 1
        if re.search(r"(nhà nước|ngoài nhà nước|đtnn)", str(a), flags=re.IGNORECASE):
            counts["scope"] += 1
    counts = sorted(counts.items(), key=lambda x: x[1] * -1)
    if counts[0][1] == 0:
        return "category"
    return counts[0][0]   

def column_data_type(arr):
    arr = list(arr)
    uni = len(set([a for a in arr if a is not None]))
    counts = {"int": 0, "float": 0, "keyword": 0, "datetime": 0, "bool": 0}
    trans = {t: f for f, t in
             [(int, "int"), (float, "float"), (trans_datatime, "datetime"), (trans_bool, "bool"), (str, "keyword")]}
    for a in arr:
        if a is None:
            continue
        if re.match(r"[+-]?[0-9]+(\.0+)?$", str(a).replace("%%", "")):
            counts["int"] += 1
        elif re.match(r"[+-]?[0-9.]+$", str(a).replace("%%", "")):
            counts["float"] += 1
        elif re.match(r"(true|yes|是|\*|✓|✔|☑|✅|√|false|no|否|⍻|×)$", str(a), flags=re.IGNORECASE):
            counts["bool"] += 1
        elif trans_datatime(str(a)):
            counts["datetime"] += 1
        else:
            counts["keyword"] += 1
    counts = sorted(counts.items(), key=lambda x: x[1] * -1)
    ty = counts[0][0]
    
    for i in range(len(arr)):
        if arr[i] is None:
            continue
        try:
            arr[i] = trans[ty](str(arr[i]))
        except Exception as e:
            arr[i] = None
    # if ty == "text":
    #    if len(arr) > 128 and uni / len(arr) < 0.1:
    #        ty = "keyword"
    return arr, ty

def chunk(filename, binary=None, from_page=0, to_page=100000000, 
          lang="Vietnamese", callback=None, **kwargs):
    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        callback(0.1, "Start to parse.")
        excel_parser = Excel()
        dfs, titles = excel_parser(
            filename,
            binary,
            from_page=0,
            to_page=1000000,
            callback=callback)
    else:
        raise NotImplementedError(
            "file type not supported yet(excel, text, csv supported)")
    res = []
    fields_map = {
        "text": "_tks",
        "int": "_long",
        "keyword": "_kwd",
        "float": "_flt",
        "datetime": "_dt",
        "bool": "_kwd"
    }
    total_fields = []
    for i, df in enumerate(dfs):
        for n in ["id", "_id", "index", "idx"]:
            if n in df.columns:
                del df[n]
        sheetname_id = get_uuid()
        clmns = df.columns.values
        txts = list(copy.deepcopy(clmns))
        py_clmns = [
            re.sub(
                r"(/.*|（[^（）]+?）|\([^()]+?\))",
                "",
                str(n)).replace(" ", "_").lower() for n in clmns
        ]
        clmn_tys = []
        for j in range(len(clmns)):
            cln, ty = column_data_type(df[clmns[j]])
            clmn_tys.append(ty)
            df[clmns[j]] = cln
            if ty == "keyword":
                txts.extend([str(c) for c in cln if c])
        clmns_map = [(py_clmns[i].lower() + fields_map[clmn_tys[i]], str(clmns[i]).replace("_", " "))
                     for i in range(len(clmns))]

        eng = lang.lower() == "english"  # is_english(txts)

        for ii, row in df.iterrows():
            d = {
                "docnm_kwd": filename,
                "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename)),
                "title_kwd": filename,
                "field_maps_tks": json.dumps({**{adjust_header(k): v for k, v in clmns_map}}, ensure_ascii=False),
                "sheetname_id": sheetname_id
            }
            if len(clmns) > 70:
                continue
            row_txt = []

            for j in range(len(clmns)):
                if row[clmns[j]] is None:
                    continue
                if not str(row[clmns[j]]):
                    continue
                if pd.isna(row[clmns[j]]):
                    continue
                fld = adjust_header(clmns_map[j][0])
                d[fld] = row[clmns[j]] if clmn_tys[j] != "keyword" else rag_tokenizer.tokenize(
                    row[clmns[j]]).replace("_", " ")
                row_txt.append("{}:{}".format(clmns[j], row[clmns[j]]))
            if not row_txt:
                continue
            tokenize(d, str(titles[i]) + ": " + str("; ".join(row_txt)), eng)
            res.append(d)
    callback(0.35, "")
    
    return res
        
        
if __name__ == "__main__":
    def donothing(progress, msg):
        pass
    from functools import partial
    callback = partial(donothing)
    res = chunk("./gso.xlsx", callback=callback)
    # print(len(res))