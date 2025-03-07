import copy
from io import BytesIO
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from services.deepdoc.parser import ExcelParser
from dateutil.parser import parse as datetime_parse
from services.nlp import rag_tokenizer, tokenize
from db.db_services.knowledgebase_service import KnowledgebaseService
class Excel(ExcelParser):
    def __call__(self, fnm, binary=None, from_page=0,
                 to_page=1000000000, callback=None):
        if not binary:
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))
            
        res, fails, done = [], [], 0
        rn = 0
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows: continue
            headers = [cell.value for cell in rows[0]]
            missed = set([i for i, h in enumerate(headers) if h is None])
            headers = [
                cell.value for i,
                cell in enumerate(
                    rows[0]) if i not in missed]
            if not headers:continue
            data = []
            for i, r in enumerate(rows[1:]):
                rn += 1
                if rn - 1 < from_page:
                    continue
                if rn - 1 >= to_page:
                    break
                row = [
                    cell.value for ii,
                    cell in enumerate(r) if ii not in missed]
                if len(row) != len(headers):
                    fails.append(str(i))
                    continue
                data.append(row)
                done += 1
            res.append(pd.DataFrame(np.array(data), columns=headers))

        callback(0.3, ("Extract records: {}~{}".format(from_page + 1, min(to_page, from_page + rn)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        return res
    
def trans_datatime(s):
    try:
        return datetime_parse(s.strip()).strftime("%Y-%m-%d %H:%M:%S")
    except Exception as e:
        pass

def trans_bool(s):
    if re.match(r"(đúng|true|yes|是|\*|✓|✔|☑|✅|√)$",
                str(s).strip(), flags=re.IGNORECASE):
        return "yes"
    if re.match(r"(không|sai|false|no|否|⍻|×)$", str(s).strip(), flags=re.IGNORECASE):
        return "no"

def column_data_type(arr):
    arr = list(arr)
    uni = len(set([a for a in arr if a is not None]))
    counts = {"int": 0, "float": 0, "text": 0, "datetime": 0, "bool": 0}
    trans = {t: f for f, t in
             [(int, "int"), (float, "float"), (trans_datatime, "datetime"), (trans_bool, "bool"), (str, "text")]}
    for a in arr:
        if a is None:
            continue
        if re.match(r"[+-]?[0-9]+(\.0+)?$", str(a).replace("%%", "")):
            counts["int"] += 1
        elif re.match(r"[+-]?[0-9.]+$", str(a).replace("%%", "")):
            counts["float"] += 1
        elif re.match(r"(true|yes|是|\*|✓|✔|☑|✅|√|false|no|否|⍻|×)$", str(a), flags=re.IGNORECASE):
            counts["bool"] += 1
        elif trans_datatime(str(a)):
            counts["datetime"] += 1
        else:
            counts["text"] += 1
    counts = sorted(counts.items(), key=lambda x: x[1] * -1)
    ty = counts[0][0]
    for i in range(len(arr)):
        if arr[i] is None:
            continue
        try:
            arr[i] = trans[ty](str(arr[i]))
        except Exception as e:
            arr[i] = None
    # if ty == "text":
    #    if len(arr) > 128 and uni / len(arr) < 0.1:
    #        ty = "keyword"
    return arr, ty

def chunk(filename, binary=None, from_page=0, to_page=100000000, 
          lang="Vietnamese", callback=None, **kwargs):
    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        callback(0.1, "Start to parse.")
        excel_parser = Excel()
        dfs = excel_parser(
            filename,
            binary,
            from_page=from_page,
            to_page=to_page,
            callback=callback)
    else:
        raise NotImplementedError(
            "file type not supported yet(excel, text, csv supported)")
    res = []
    fields_map = {
        "text": "_tks",
        "int": "_long",
        "keyword": "_kwd",
        "float": "_flt",
        "datetime": "_dt",
        "bool": "_kwd"
    }
    for df in dfs:
        for n in ["id", "_id", "index", "idx"]:
            if n in df.columns:
                del df[n]
        clmns = df.columns.values
        txts = list(copy.deepcopy(clmns))
        py_clmns = [
            re.sub(
                r"(/.*|（[^（）]+?）|\([^()]+?\))",
                "",
                str(n)).replace(" ", "_").lower() for n in clmns
        ]
        clmn_tys = []
        for j in range(len(clmns)):
            cln, ty = column_data_type(df[clmns[j]])
            clmn_tys.append(ty)
            df[clmns[j]] = cln
            if ty == "text":
                txts.extend([str(c) for c in cln if c])
        clmns_map = [(py_clmns[i].lower() + fields_map[clmn_tys[i]], str(clmns[i]).replace("_", " "))
                     for i in range(len(clmns))]

        eng = lang.lower() == "english"  # is_english(txts)
        for ii, row in df.iterrows():
            d = {
                "docnm_kwd": filename,
                "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename)),
                "title_kwd": filename
            }
            row_txt = []
            for j in range(len(clmns)):
                if row[clmns[j]] is None:
                    continue
                if not str(row[clmns[j]]):
                    continue
                if pd.isna(row[clmns[j]]):
                    continue
                fld = clmns_map[j][0]
                d[fld] = row[clmns[j]] if clmn_tys[j] != "text" else rag_tokenizer.tokenize(
                    row[clmns[j]]).replace("_", " ")
                row_txt.append("{}:{}".format(clmns[j], row[clmns[j]]))
            if not row_txt:
                continue
            tokenize(d, "; ".join(row_txt), eng)
            res.append(d)
        KnowledgebaseService.update_parser_config(
            kwargs["kb_id"], {"field_map": {k: v for k, v in clmns_map}})
    callback(0.35, "")
    
    return res
        