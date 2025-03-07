from io import BytesIO
from copy import deepcopy
import re
from openpyxl import load_workbook
from services.deepdoc.parser import ExcelParser
from services.nlp import rag_tokenizer, tokenize
class Excel(ExcelParser):
    def __call__(self, fnm, binary=None, callback=None):
        if not binary:
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))

        res, fails = [], []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            for i, r in enumerate(rows):
                if i == 0:
                    continue
                q, a = "", ""
                for cell in r:
                    if not cell.value:
                        continue
                    if not q:
                        q = str(cell.value)
                    elif not a:
                        a = str(cell.value)
                    else:
                        break
                if q and a:
                    res.append((q, a))
                else:
                    fails.append(str(i + 1))
                if len(res) % 999 == 0:
                    callback(len(res) *
                             0.6 /
                             total, ("Extract Q&A: {}".format(len(res)) +
                                     (f"{len(fails)} failure, line: %s..." %
                                      (",".join(fails[:3])) if fails else "")))

        callback(0.6, ("Extract Q&A: {}. ".format(len(res)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        return res

def rmPrefix(txt):
    return re.sub(
        r"^(问题|答案|回答|user|assistant|Q|A|Question|Answer|问|答|Câu hỏi|Câu trả lời)[\t:： ]+", "", txt.strip(), flags=re.IGNORECASE)


def beAdoc(d, q, a, eng):
    qprefix = "Question: " if eng else "Câu hỏi："
    aprefix = "Answer: " if eng else "Câu trả lời："
    d["content_with_weight"] = "\t".join(
        [qprefix + rmPrefix(q), aprefix + rmPrefix(a)])
    d["content_ltks"] = rag_tokenizer.tokenize(q)
    d["content_sm_ltks"] = rag_tokenizer.fine_grained_tokenize(d["content_ltks"])
    return d
    
def chunk(filename, binary=None, lang="Vietnamese", callback=None, **kwargs):
    eng = False
    res = []
    doc = {
        "docnm_kwd": filename,
        "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename)),
        "title_kwd": filename
    }
    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        callback(0.1, "Start to parse.")
        excel_parser = Excel()
        for q, a in excel_parser(filename, binary, callback):
            res.append(beAdoc(deepcopy(doc), q, a, eng))
        return res
    else:
        raise NotImplementedError(
            "file type not supported yet(excel, text, csv supported)")
        
if __name__ == "__main__":
    def donothing(progress, msg):
        pass
    from functools import partial
    callback = partial(donothing)
    res = chunk("./qa_test.xlsx", callback=callback)
    print(res)